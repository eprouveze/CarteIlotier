#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Répartiteur de Zones d'Urgence - Ilotiers
Application pour répartir équitablement les familles entre deux ilotiers
Version 1.0 - 2025
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import csv
import json
import os
import sys
import time
import threading
from datetime import datetime
from typing import List, Dict, Tuple
import math

# Try to import required libraries
try:
    import googlemaps
    GOOGLEMAPS_AVAILABLE = True
except ImportError:
    GOOGLEMAPS_AVAILABLE = False

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

class IlotierZoneMapper:
    def __init__(self, root):
        self.root = root
        self.root.title("Répartiteur de Zones d'Urgence - Ilotiers")
        self.root.geometry("900x750")
        
        # Variables
        self.csv_file = tk.StringVar()
        self.api_key = tk.StringVar()
        self.ilotier1_name = tk.StringVar(value="Ilotier 1")
        self.ilotier1_address = tk.StringVar()
        self.ilotier2_name = tk.StringVar(value="Ilotier 2")
        self.ilotier2_address = tk.StringVar()
        self.processing = False
        
        # Load saved API key if exists
        self.load_api_key()
        
        # Create UI
        self.create_widgets()
        
    def create_widgets(self):
        """Créer l'interface utilisateur"""
        
        # Style
        style = ttk.Style()
        style.configure('Header.TLabel', font=('Helvetica', 14, 'bold'))
        style.configure('Section.TLabel', font=('Helvetica', 12, 'bold'))
        
        # Main container with padding
        main_frame = ttk.Frame(self.root, padding="20")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Configure grid weights
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)
        
        row = 0
        
        # Title
        title_label = ttk.Label(main_frame, text="Répartiteur de Zones d'Urgence", style='Header.TLabel')
        title_label.grid(row=row, column=0, columnspan=3, pady=(0, 20))
        row += 1
        
        # Separator
        ttk.Separator(main_frame, orient='horizontal').grid(row=row, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=10)
        row += 1
        
        # Section 1: Fichier CSV
        section1_label = ttk.Label(main_frame, text="1. Fichier de données", style='Section.TLabel')
        section1_label.grid(row=row, column=0, sticky=tk.W, pady=(10, 5))
        row += 1
        
        ttk.Label(main_frame, text="Fichier de données :").grid(row=row, column=0, sticky=tk.W, padx=(20, 0))
        ttk.Entry(main_frame, textvariable=self.csv_file, width=40).grid(row=row, column=1, sticky=(tk.W, tk.E), padx=5)
        ttk.Button(main_frame, text="Parcourir...", command=self.browse_csv).grid(row=row, column=2)
        row += 1
        
        csv_help = ttk.Label(main_frame, text="Format accepté : Excel (.xlsx) ou CSV exporté du registre consulaire", 
                            font=('Helvetica', 9), foreground='gray')
        csv_help.grid(row=row, column=1, sticky=tk.W, padx=5, pady=(0, 10))
        row += 1
        
        # Security notice
        security_frame = ttk.Frame(main_frame)
        security_frame.grid(row=row, column=0, columnspan=3, sticky=(tk.W, tk.E), padx=(20, 0), pady=(0, 10))
        
        security_icon = ttk.Label(security_frame, text="🔒", font=('Helvetica', 14))
        security_icon.pack(side=tk.LEFT, padx=(0, 5))
        
        security_label = ttk.Label(security_frame, 
                                  text="Sécurité : Seuls les codes famille seront affichés sur la carte (aucun nom ni information personnelle)",
                                  font=('Helvetica', 10, 'bold'), foreground='green')
        security_label.pack(side=tk.LEFT)
        row += 1
        
        # Separator
        ttk.Separator(main_frame, orient='horizontal').grid(row=row, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=10)
        row += 1
        
        # Section 2: Ilotiers
        section2_label = ttk.Label(main_frame, text="2. Informations des ilotiers", style='Section.TLabel')
        section2_label.grid(row=row, column=0, sticky=tk.W, pady=(10, 5))
        row += 1
        
        # Ilotier 1
        ilotier1_frame = ttk.LabelFrame(main_frame, text="Ilotier 1", padding="10")
        ilotier1_frame.grid(row=row, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=5, padx=(20, 0))
        
        ttk.Label(ilotier1_frame, text="Nom :").grid(row=0, column=0, sticky=tk.W)
        ttk.Entry(ilotier1_frame, textvariable=self.ilotier1_name, width=30).grid(row=0, column=1, sticky=(tk.W, tk.E), padx=5)
        
        ttk.Label(ilotier1_frame, text="Adresse :").grid(row=1, column=0, sticky=tk.W, pady=(5, 0))
        ttk.Entry(ilotier1_frame, textvariable=self.ilotier1_address, width=50).grid(row=1, column=1, sticky=(tk.W, tk.E), padx=5, pady=(5, 0))
        
        ilotier1_frame.columnconfigure(1, weight=1)
        row += 1
        
        # Ilotier 2
        ilotier2_frame = ttk.LabelFrame(main_frame, text="Ilotier 2", padding="10")
        ilotier2_frame.grid(row=row, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=5, padx=(20, 0))
        
        ttk.Label(ilotier2_frame, text="Nom :").grid(row=0, column=0, sticky=tk.W)
        ttk.Entry(ilotier2_frame, textvariable=self.ilotier2_name, width=30).grid(row=0, column=1, sticky=(tk.W, tk.E), padx=5)
        
        ttk.Label(ilotier2_frame, text="Adresse :").grid(row=1, column=0, sticky=tk.W, pady=(5, 0))
        ttk.Entry(ilotier2_frame, textvariable=self.ilotier2_address, width=50).grid(row=1, column=1, sticky=(tk.W, tk.E), padx=5, pady=(5, 0))
        
        ilotier2_frame.columnconfigure(1, weight=1)
        row += 1
        
        # Separator
        ttk.Separator(main_frame, orient='horizontal').grid(row=row, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=10)
        row += 1
        
        # Section 3: Clé API Google Maps
        section3_label = ttk.Label(main_frame, text="3. Clé API Google Maps", style='Section.TLabel')
        section3_label.grid(row=row, column=0, sticky=tk.W, pady=(10, 5))
        row += 1
        
        ttk.Label(main_frame, text="Clé API :").grid(row=row, column=0, sticky=tk.W, padx=(20, 0))
        api_entry = ttk.Entry(main_frame, textvariable=self.api_key, width=40, show="*")
        api_entry.grid(row=row, column=1, sticky=(tk.W, tk.E), padx=5)
        ttk.Button(main_frame, text="Sauvegarder", command=self.save_api_key).grid(row=row, column=2)
        row += 1
        
        api_help_frame = ttk.Frame(main_frame)
        api_help_frame.grid(row=row, column=1, columnspan=2, sticky=tk.W, padx=5, pady=(0, 10))
        
        api_help1 = ttk.Label(api_help_frame, text="Pas de clé API ? ", font=('Helvetica', 9), foreground='gray')
        api_help1.pack(side=tk.LEFT)
        
        api_help_link = ttk.Label(api_help_frame, text="Voir le guide", font=('Helvetica', 9, 'underline'), 
                                 foreground='blue', cursor='hand2')
        api_help_link.pack(side=tk.LEFT)
        api_help_link.bind("<Button-1>", lambda e: self.show_api_guide())
        
        api_help2 = ttk.Label(api_help_frame, text=" ou contactez Emmanuel", font=('Helvetica', 9), foreground='gray')
        api_help2.pack(side=tk.LEFT)
        row += 1
        
        # Separator
        ttk.Separator(main_frame, orient='horizontal').grid(row=row, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=10)
        row += 1
        
        # Process button
        self.process_btn = ttk.Button(main_frame, text="Lancer le traitement", command=self.process_data)
        self.process_btn.grid(row=row, column=0, columnspan=3, pady=20)
        row += 1
        
        # Progress bar
        self.progress = ttk.Progressbar(main_frame, mode='indeterminate')
        self.progress.grid(row=row, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 10))
        self.progress.grid_remove()  # Hidden initially
        row += 1
        
        # Status/Log area
        log_label = ttk.Label(main_frame, text="Journal d'exécution :", style='Section.TLabel')
        log_label.grid(row=row, column=0, sticky=tk.W, pady=(10, 5))
        row += 1
        
        self.log_text = scrolledtext.ScrolledText(main_frame, height=10, width=80, wrap=tk.WORD)
        self.log_text.grid(row=row, column=0, columnspan=3, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(0, 10))
        main_frame.rowconfigure(row, weight=1)
        row += 1
        
        # Footer
        footer_frame = ttk.Frame(main_frame)
        footer_frame.grid(row=row, column=0, columnspan=3, sticky=(tk.W, tk.E))
        
        ttk.Label(footer_frame, text="Version 1.0 - 2025", font=('Helvetica', 9), foreground='gray').pack(side=tk.LEFT)
        
        help_btn = ttk.Button(footer_frame, text="Aide", command=self.show_help)
        help_btn.pack(side=tk.RIGHT, padx=(5, 0))
        
        about_btn = ttk.Button(footer_frame, text="À propos", command=self.show_about)
        about_btn.pack(side=tk.RIGHT)
        
    def browse_csv(self):
        """Sélectionner un fichier CSV ou Excel"""
        filename = filedialog.askopenfilename(
            title="Sélectionner le fichier de données",
            filetypes=[
                ("Fichiers Excel", "*.xlsx;*.xls"),
                ("Fichiers CSV", "*.csv"),
                ("Tous les fichiers", "*.*")
            ]
        )
        if filename:
            self.csv_file.set(filename)
            self.log(f"Fichier sélectionné : {os.path.basename(filename)}")
            
    def log(self, message):
        """Ajouter un message au journal"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.log_text.insert(tk.END, f"[{timestamp}] {message}\n")
        self.log_text.see(tk.END)
        self.root.update()
        
    def save_api_key(self):
        """Sauvegarder la clé API"""
        api_key = self.api_key.get().strip()
        if api_key:
            config_file = os.path.expanduser("~/.ilotier_config.json")
            try:
                with open(config_file, 'w') as f:
                    json.dump({"api_key": api_key}, f)
                self.log("Clé API sauvegardée avec succès")
                messagebox.showinfo("Succès", "Clé API sauvegardée")
            except Exception as e:
                self.log(f"Erreur lors de la sauvegarde : {str(e)}")
                messagebox.showerror("Erreur", f"Impossible de sauvegarder la clé : {str(e)}")
                
    def load_api_key(self):
        """Charger la clé API sauvegardée"""
        config_file = os.path.expanduser("~/.ilotier_config.json")
        if os.path.exists(config_file):
            try:
                with open(config_file, 'r') as f:
                    config = json.load(f)
                    self.api_key.set(config.get("api_key", ""))
            except:
                pass
                
    def validate_inputs(self):
        """Valider les entrées utilisateur"""
        errors = []
        
        if not self.csv_file.get():
            errors.append("- Veuillez sélectionner un fichier CSV")
            
        if not self.api_key.get():
            errors.append("- Veuillez entrer une clé API Google Maps")
            
        if not self.ilotier1_address.get():
            errors.append("- Veuillez entrer l'adresse de l'ilotier 1")
            
        if not self.ilotier2_address.get():
            errors.append("- Veuillez entrer l'adresse de l'ilotier 2")
            
        if errors:
            messagebox.showerror("Données manquantes", "Veuillez corriger les erreurs suivantes :\n\n" + "\n".join(errors))
            return False
            
        if not GOOGLEMAPS_AVAILABLE:
            messagebox.showerror("Module manquant", 
                "Le module 'googlemaps' n'est pas installé.\n\n" +
                "Veuillez l'installer avec : pip install googlemaps")
            return False
            
        return True
        
    def process_data(self):
        """Lancer le traitement dans un thread séparé"""
        if not self.validate_inputs():
            return
            
        if self.processing:
            messagebox.showwarning("Traitement en cours", "Un traitement est déjà en cours")
            return
            
        # Clear log
        self.log_text.delete(1.0, tk.END)
        
        # Start processing in a separate thread
        thread = threading.Thread(target=self.process_data_thread)
        thread.daemon = True
        thread.start()
        
    def process_data_thread(self):
        """Thread de traitement des données"""
        self.processing = True
        self.process_btn.config(state='disabled')
        self.progress.grid()
        self.progress.start()
        
        try:
            self.log("Début du traitement...")
            
            # Initialize Google Maps client
            gmaps = googlemaps.Client(key=self.api_key.get())
            
            # Load and process data
            families = self.load_families_from_csv(self.csv_file.get())
            self.log(f"Chargé {len(families)} familles du CSV")
            
            # Geocode ilotiers
            ilotier_coords = self.geocode_ilotiers(gmaps)
            
            # Geocode families
            self.geocode_families(families, gmaps)
            
            # Apply balance algorithm
            self.apply_balance_algorithm(families, ilotier_coords)
            
            # Save outputs
            self.save_outputs(families, ilotier_coords)
            
            self.log("Traitement terminé avec succès !")
            messagebox.showinfo("Succès", 
                "Le traitement est terminé !\n\n" +
                "Fichiers générés :\n" +
                "- zones.json (données pour visualisation)\n" +
                "- familles_zones.csv (CSV mis à jour)\n" +
                "- zones.kml (pour Google My Maps)")
            
        except Exception as e:
            self.log(f"ERREUR : {str(e)}")
            messagebox.showerror("Erreur", f"Une erreur est survenue :\n\n{str(e)}")
            
        finally:
            self.processing = False
            self.process_btn.config(state='normal')
            self.progress.stop()
            self.progress.grid_remove()
            
    def load_families_from_csv(self, filename):
        """Charger les familles depuis le fichier CSV ou Excel"""
        families_dict = {}
        
        # Check file extension
        file_ext = os.path.splitext(filename)[1].lower()
        
        if file_ext in ['.xlsx', '.xls']:
            # Load from Excel
            if not OPENPYXL_AVAILABLE:
                raise Exception("Le module 'openpyxl' n'est pas installé. Installez-le avec : pip install openpyxl")
                
            wb = openpyxl.load_workbook(filename, read_only=True)
            ws = wb.active
            
            # Get headers from first row
            headers = []
            for cell in ws[1]:
                headers.append(cell.value if cell.value else '')
                
            # Process rows
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                # Create dict from row
                row_dict = {}
                for col_idx, value in enumerate(row):
                    if col_idx < len(headers):
                        row_dict[headers[col_idx]] = str(value) if value is not None else ''
                        
                self.process_family_row(row_dict, families_dict)
                
            wb.close()
            
        else:
            # Load from CSV
            with open(filename, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    self.process_family_row(row, families_dict)
                    
        return list(families_dict.values())
        
    def process_family_row(self, row, families_dict):
        """Traiter une ligne de données famille"""
        family_code = row.get('Famille', '').strip()
        if not family_code or family_code == 'None':
            return
            
        # Skip invalid family codes (example for German families)
        # You can customize this list based on your needs
        if family_code in ['2165915', '2192050']:
            return
            
        # Build address
        address_parts = []
        if row.get('Adresse postale') and row.get('Adresse postale') != 'None':
            address_parts.append(row['Adresse postale'].strip())
        if row.get('Code postal de résidence') and row.get('Code postal de résidence') != 'None':
            address_parts.append(row['Code postal de résidence'].strip())
        if row.get('Ville de residence') and row.get('Ville de residence') != 'None':
            address_parts.append(row['Ville de residence'].strip())
            
        address = ', '.join(address_parts)
        
        # Get contact info
        is_main = row.get('Personne lien (O/N)', '') == 'O'
        
        # Store or update family
        if family_code not in families_dict or is_main:
            prenom = row.get('Prénoms', '')
            nom = row.get('Nom de famille', '')
            if prenom == 'None': prenom = ''
            if nom == 'None': nom = ''
            
            families_dict[family_code] = {
                'family_code': family_code,
                'address': address,
                'contact_name': f"{prenom} {nom}".strip(),
                'mobile': row.get('Mobile perso', '').strip() if row.get('Mobile perso') != 'None' else '',
                'phone': row.get('Téléphone perso', '').strip() if row.get('Téléphone perso') != 'None' else '',
                'email': row.get('Adresse électronique 2 (Usage communication consulaire et vote électronique)', '').strip() if row.get('Adresse électronique 2 (Usage communication consulaire et vote électronique)') != 'None' else '',
                'raw_data': row
            }
        
    def geocode_ilotiers(self, gmaps):
        """Géocoder les adresses des ilotiers"""
        self.log("Géocodage des ilotiers...")
        
        ilotiers = {
            self.ilotier1_name.get(): {
                'address': self.ilotier1_address.get(),
                'zone': 1
            },
            self.ilotier2_name.get(): {
                'address': self.ilotier2_address.get(),
                'zone': 2
            }
        }
        
        for name, info in ilotiers.items():
            try:
                result = gmaps.geocode(info['address'])
                if result:
                    location = result[0]['geometry']['location']
                    info['lat'] = location['lat']
                    info['lng'] = location['lng']
                    self.log(f"  {name} : {info['lat']:.6f}, {info['lng']:.6f}")
                else:
                    raise Exception(f"Impossible de géocoder l'adresse de {name}")
            except Exception as e:
                raise Exception(f"Erreur lors du géocodage de {name}: {str(e)}")
                
        return ilotiers
        
    def geocode_families(self, families, gmaps):
        """Géocoder les adresses des familles"""
        self.log(f"Géocodage de {len(families)} familles...")
        
        for i, family in enumerate(families):
            if i % 10 == 0:
                self.log(f"  Progression : {i}/{len(families)}")
                
            try:
                result = gmaps.geocode(family['address'])
                if result:
                    location = result[0]['geometry']['location']
                    family['lat'] = location['lat']
                    family['lng'] = location['lng']
                else:
                    family['lat'] = None
                    family['lng'] = None
                    self.log(f"  ⚠ Impossible de géocoder : {family['family_code']}")
            except Exception as e:
                family['lat'] = None
                family['lng'] = None
                self.log(f"  ⚠ Erreur pour {family['family_code']}: {str(e)}")
                
            # Rate limiting
            time.sleep(0.2)
            
        self.log(f"  Géocodage terminé")
        
    def calculate_distance(self, lat1, lng1, lat2, lng2):
        """Calculer la distance entre deux points (formule Haversine)"""
        R = 6371  # Rayon de la Terre en km
        
        phi1 = math.radians(lat1)
        phi2 = math.radians(lat2)
        delta_phi = math.radians(lat2 - lat1)
        delta_lambda = math.radians(lng2 - lng1)
        
        a = math.sin(delta_phi/2)**2 + math.cos(phi1) * math.cos(phi2) * math.sin(delta_lambda/2)**2
        c = 2 * math.atan2(math.sqrt(a), math.sqrt(1-a))
        
        return R * c
        
    def apply_balance_algorithm(self, families, ilotier_coords):
        """Appliquer l'algorithme d'équilibrage des zones"""
        self.log("Application de l'algorithme d'équilibrage...")
        
        # Get ilotier names
        ilotier_names = list(ilotier_coords.keys())
        ilotier1_name = ilotier_names[0]
        ilotier2_name = ilotier_names[1]
        
        # Calculate distances for each family
        valid_families = []
        for family in families:
            if family.get('lat') is None:
                continue
                
            family['dist_ilotier1'] = self.calculate_distance(
                family['lat'], family['lng'],
                ilotier_coords[ilotier1_name]['lat'], 
                ilotier_coords[ilotier1_name]['lng']
            )
            
            family['dist_ilotier2'] = self.calculate_distance(
                family['lat'], family['lng'],
                ilotier_coords[ilotier2_name]['lat'],
                ilotier_coords[ilotier2_name]['lng']
            )
            
            valid_families.append(family)
            
        # Natural Voronoi assignment
        for family in valid_families:
            if family['dist_ilotier1'] <= family['dist_ilotier2']:
                family['natural_zone'] = 1
            else:
                family['natural_zone'] = 2
                
        # Count natural distribution
        zone1_count = sum(1 for f in valid_families if f['natural_zone'] == 1)
        zone2_count = sum(1 for f in valid_families if f['natural_zone'] == 2)
        
        self.log(f"  Répartition naturelle : Zone 1={zone1_count}, Zone 2={zone2_count}")
        
        # Balance if needed
        total = len(valid_families)
        target = total // 2
        imbalance = zone1_count - zone2_count
        
        if abs(imbalance) > 2:
            # Need rebalancing
            self.log(f"  Rééquilibrage nécessaire...")
            
            # Calculate transferability
            for family in valid_families:
                dist_diff = abs(family['dist_ilotier1'] - family['dist_ilotier2'])
                max_dist = max(family['dist_ilotier1'], family['dist_ilotier2'])
                family['transferability'] = dist_diff / max_dist if max_dist > 0 else 0
                
            # Sort by transferability
            valid_families.sort(key=lambda x: x['transferability'])
            
            # Transfer families to achieve balance
            transferred = 0
            need_to_transfer = abs(imbalance) // 2
            
            for family in valid_families:
                if transferred >= need_to_transfer:
                    break
                    
                if imbalance > 0 and family['natural_zone'] == 1:
                    # Transfer from zone 1 to zone 2
                    family['zone'] = 2
                    family['ilotier'] = ilotier2_name
                    family['distance'] = family['dist_ilotier2']
                    transferred += 1
                elif imbalance < 0 and family['natural_zone'] == 2:
                    # Transfer from zone 2 to zone 1
                    family['zone'] = 1
                    family['ilotier'] = ilotier1_name
                    family['distance'] = family['dist_ilotier1']
                    transferred += 1
                else:
                    # Keep natural assignment
                    family['zone'] = family['natural_zone']
                    if family['zone'] == 1:
                        family['ilotier'] = ilotier1_name
                        family['distance'] = family['dist_ilotier1']
                    else:
                        family['ilotier'] = ilotier2_name
                        family['distance'] = family['dist_ilotier2']
                        
            # Assign remaining families
            for family in valid_families:
                if 'zone' not in family:
                    family['zone'] = family['natural_zone']
                    if family['zone'] == 1:
                        family['ilotier'] = ilotier1_name
                        family['distance'] = family['dist_ilotier1']
                    else:
                        family['ilotier'] = ilotier2_name
                        family['distance'] = family['dist_ilotier2']
                        
            self.log(f"  {transferred} familles transférées pour l'équilibrage")
        else:
            # Keep natural assignment
            for family in valid_families:
                family['zone'] = family['natural_zone']
                if family['zone'] == 1:
                    family['ilotier'] = ilotier1_name
                    family['distance'] = family['dist_ilotier1']
                else:
                    family['ilotier'] = ilotier2_name
                    family['distance'] = family['dist_ilotier2']
                    
        # Final count
        zone1_final = sum(1 for f in valid_families if f['zone'] == 1)
        zone2_final = sum(1 for f in valid_families if f['zone'] == 2)
        
        self.log(f"  Répartition finale : Zone 1={zone1_final}, Zone 2={zone2_final}")
        self.log(f"  Équilibre : {zone1_final/total*100:.1f}% / {zone2_final/total*100:.1f}%")
        
    def save_outputs(self, families, ilotier_coords):
        """Sauvegarder les fichiers de sortie"""
        output_dir = os.path.dirname(self.csv_file.get())
        
        # Save JSON
        json_file = os.path.join(output_dir, "zones.json")
        with open(json_file, 'w', encoding='utf-8') as f:
            json.dump({
                'ilotiers': ilotier_coords,
                'families': families
            }, f, ensure_ascii=False, indent=2)
        self.log(f"Fichier JSON sauvegardé : zones.json")
        
        # Save updated CSV
        csv_file = os.path.join(output_dir, "familles_zones.csv")
        self.save_updated_csv(families, csv_file)
        self.log(f"Fichier CSV sauvegardé : familles_zones.csv")
        
        # Save KML
        kml_file = os.path.join(output_dir, "zones.kml")
        self.save_kml(families, ilotier_coords, kml_file)
        self.log(f"Fichier KML sauvegardé : zones.kml")
        
    def save_updated_csv(self, families, output_file):
        """Sauvegarder le CSV mis à jour"""
        # Read original CSV to preserve all columns
        original_rows = []
        with open(self.csv_file.get(), 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            fieldnames = reader.fieldnames
            for row in reader:
                original_rows.append(row)
                
        # Add new columns
        new_fieldnames = fieldnames + ['Ilotier', 'Zone', 'Distance (km)', 'Lien Google Maps']
        
        # Create family lookup
        family_lookup = {f['family_code']: f for f in families if 'zone' in f}
        
        # Update rows
        for row in original_rows:
            family_code = row.get('Famille', '')
            if family_code in family_lookup:
                family = family_lookup[family_code]
                row['Ilotier'] = family.get('ilotier', '')
                row['Zone'] = str(family.get('zone', ''))
                row['Distance (km)'] = f"{family.get('distance', 0):.2f}"
                row['Lien Google Maps'] = f"https://www.google.com/maps/search/?api=1&query={family.get('lat', '')},{family.get('lng', '')}"
            else:
                row['Ilotier'] = ''
                row['Zone'] = ''
                row['Distance (km)'] = ''
                row['Lien Google Maps'] = ''
                
        # Write updated CSV
        with open(output_file, 'w', encoding='utf-8', newline='') as f:
            writer = csv.DictWriter(f, fieldnames=new_fieldnames)
            writer.writeheader()
            writer.writerows(original_rows)
            
    def save_kml(self, families, ilotier_coords, output_file):
        """Générer le fichier KML pour Google My Maps"""
        kml_content = '''<?xml version="1.0" encoding="UTF-8"?>
<kml xmlns="http://www.opengis.net/kml/2.2">
    <Document>
        <name>Zones d'urgence</name>
        <description>Répartition des familles entre ilotiers</description>
        
        <!-- Style Zone 1 -->
        <Style id="zone1">
            <IconStyle>
                <color>ffff0000</color>
                <scale>0.8</scale>
                <Icon>
                    <href>http://maps.google.com/mapfiles/kml/paddle/blu-circle.png</href>
                </Icon>
            </IconStyle>
        </Style>
        
        <!-- Style Zone 2 -->
        <Style id="zone2">
            <IconStyle>
                <color>ff0000ff</color>
                <scale>0.8</scale>
                <Icon>
                    <href>http://maps.google.com/mapfiles/kml/paddle/red-circle.png</href>
                </Icon>
            </IconStyle>
        </Style>
        
        <!-- Style Ilotier -->
        <Style id="ilotier">
            <IconStyle>
                <color>ff00ff00</color>
                <scale>1.2</scale>
                <Icon>
                    <href>http://maps.google.com/mapfiles/kml/paddle/grn-stars.png</href>
                </Icon>
            </IconStyle>
        </Style>
        
        <!-- Ilotiers -->
        <Folder>
            <name>Ilotiers</name>
'''
        
        # Add ilotiers
        for name, coords in ilotier_coords.items():
            kml_content += f'''            <Placemark>
                <name>{name}</name>
                <description>Ilotier Zone {coords['zone']}</description>
                <styleUrl>#ilotier</styleUrl>
                <Point>
                    <coordinates>{coords['lng']},{coords['lat']},0</coordinates>
                </Point>
            </Placemark>
'''
        
        kml_content += '''        </Folder>
        
        <!-- Familles Zone 1 -->
        <Folder>
            <name>Zone 1</name>
'''
        
        # Add Zone 1 families
        for family in families:
            if family.get('zone') == 1 and family.get('lat'):
                kml_content += f'''            <Placemark>
                <name>{family['family_code']}</name>
                <description>{family.get('address', '')}
Distance : {family.get('distance', 0):.2f} km</description>
                <styleUrl>#zone1</styleUrl>
                <Point>
                    <coordinates>{family['lng']},{family['lat']},0</coordinates>
                </Point>
            </Placemark>
'''
        
        kml_content += '''        </Folder>
        
        <!-- Familles Zone 2 -->
        <Folder>
            <name>Zone 2</name>
'''
        
        # Add Zone 2 families
        for family in families:
            if family.get('zone') == 2 and family.get('lat'):
                kml_content += f'''            <Placemark>
                <name>{family['family_code']}</name>
                <description>{family.get('address', '')}
Distance : {family.get('distance', 0):.2f} km</description>
                <styleUrl>#zone2</styleUrl>
                <Point>
                    <coordinates>{family['lng']},{family['lat']},0</coordinates>
                </Point>
            </Placemark>
'''
        
        kml_content += '''        </Folder>
    </Document>
</kml>'''
        
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write(kml_content)
            
    def show_api_guide(self):
        """Afficher le guide pour obtenir une clé API"""
        guide_window = tk.Toplevel(self.root)
        guide_window.title("Guide - Obtenir une clé API Google Maps")
        guide_window.geometry("700x600")
        
        text = scrolledtext.ScrolledText(guide_window, wrap=tk.WORD, padx=10, pady=10)
        text.pack(fill=tk.BOTH, expand=True)
        
        guide_text = """COMMENT OBTENIR UNE CLÉ API GOOGLE MAPS

1. CRÉER UN COMPTE GOOGLE CLOUD
   • Allez sur : https://console.cloud.google.com/
   • Connectez-vous avec votre compte Google
   • Acceptez les conditions d'utilisation

2. CRÉER UN NOUVEAU PROJET
   • Cliquez sur "Sélectionner un projet" en haut
   • Cliquez sur "Nouveau projet"
   • Donnez un nom (ex: "Ilotier Zones")
   • Cliquez sur "Créer"

3. ACTIVER LES APIs NÉCESSAIRES
   • Dans le menu, allez dans "APIs et services" > "Bibliothèque"
   • Recherchez et activez :
     - Maps JavaScript API
     - Geocoding API
   • Cliquez sur "Activer" pour chaque API

4. CRÉER UNE CLÉ API
   • Allez dans "APIs et services" > "Identifiants"
   • Cliquez sur "+ Créer des identifiants" > "Clé API"
   • La clé sera créée et affichée
   • COPIEZ cette clé immédiatement !

5. SÉCURISER LA CLÉ (Recommandé)
   • Cliquez sur la clé créée
   • Dans "Restrictions d'application", sélectionnez "Adresses IP"
   • Dans "Restrictions d'API", sélectionnez "Restreindre la clé"
   • Cochez uniquement :
     - Maps JavaScript API
     - Geocoding API
   • Cliquez sur "Enregistrer"

6. FACTURATION
   • Google offre 200$ de crédit gratuit par mois
   • Pour notre usage (quelques centaines d'adresses), 
     cela sera largement suffisant et GRATUIT
   • Vous devez quand même ajouter un compte de facturation
     (carte bancaire) mais vous ne serez pas facturé

ALTERNATIVE : DEMANDER À EMMANUEL
Si vous préférez ne pas créer votre propre clé, 
vous pouvez contacter Emmanuel qui partagera la sienne :
manu@prouveze.fr

Note : La clé API est nécessaire pour géocoder les adresses
(convertir une adresse en coordonnées GPS).
"""
        
        text.insert(tk.END, guide_text)
        text.config(state=tk.DISABLED)
        
        close_btn = ttk.Button(guide_window, text="Fermer", command=guide_window.destroy)
        close_btn.pack(pady=10)
        
    def show_help(self):
        """Afficher l'aide"""
        help_window = tk.Toplevel(self.root)
        help_window.title("Aide")
        help_window.geometry("600x500")
        
        text = scrolledtext.ScrolledText(help_window, wrap=tk.WORD, padx=10, pady=10)
        text.pack(fill=tk.BOTH, expand=True)
        
        help_text = """AIDE - RÉPARTITEUR DE ZONES D'URGENCE

UTILISATION :

1. PRÉPARER LE FICHIER CSV
   • Exportez la liste des familles depuis le registre consulaire
   • Le fichier doit être au format CSV
   • Les colonnes importantes sont :
     - Famille (code famille)
     - Adresse postale
     - Code postal de résidence
     - Ville de residence
     - Informations de contact

2. CONFIGURER LES ILOTIERS
   • Entrez le nom de chaque ilotier
   • Entrez l'adresse complète de chaque ilotier
   • Ces adresses serviront de points de référence

3. ENTRER LA CLÉ API
   • Entrez votre clé API Google Maps
   • Cliquez sur "Sauvegarder" pour la mémoriser
   • Voir le guide si vous n'avez pas de clé

4. LANCER LE TRAITEMENT
   • Cliquez sur "Lancer le traitement"
   • Le processus peut prendre plusieurs minutes
   • Ne fermez pas l'application pendant le traitement

FICHIERS GÉNÉRÉS :

• zones.json : Données complètes en format JSON
• familles_zones.csv : CSV original avec colonnes ajoutées
  - Ilotier assigné
  - Numéro de zone
  - Distance en km
  - Lien Google Maps
• zones.kml : Pour import dans Google My Maps

ALGORITHME :

L'application utilise un algorithme de répartition équilibrée :
1. Assigne chaque famille à l'ilotier le plus proche
2. Si le déséquilibre est trop important (>2 familles)
3. Transfère les familles les plus "transférables"
4. Objectif : répartition 50/50

DÉPANNAGE :

• "Module googlemaps non installé"
  → Ouvrez Terminal et tapez : pip install googlemaps

• "Erreur de géocodage"
  → Vérifiez votre clé API
  → Vérifiez votre connexion internet

• "Adresse non trouvée"
  → Certaines adresses peuvent être mal formatées
  → Vérifiez le CSV original

Pour plus d'aide, contactez Emmanuel : manu@prouveze.fr
"""
        
        text.insert(tk.END, help_text)
        text.config(state=tk.DISABLED)
        
        close_btn = ttk.Button(help_window, text="Fermer", command=help_window.destroy)
        close_btn.pack(pady=10)
        
    def show_about(self):
        """Afficher À propos"""
        about_text = """Répartiteur de Zones d'Urgence
Version 1.0 - 2025

Développé pour les ilotiers du réseau consulaire
Pour répartir équitablement les familles en cas d'urgence

Créé par Emmanuel Prouvèze
manu@prouveze.fr

Cette application utilise :
• Google Maps API pour le géocodage
• Algorithme de Voronoi pondéré pour la répartition

Licence : Usage libre pour les ilotiers"""
        
        messagebox.showinfo("À propos", about_text)

def main():
    """Point d'entrée principal"""
    root = tk.Tk()
    app = IlotierZoneMapper(root)
    root.mainloop()

if __name__ == "__main__":
    main()